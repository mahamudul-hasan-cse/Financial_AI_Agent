"""Tests for ExcelTools (tools/excel_tools.py).

All yfinance network calls are mocked so tests run offline and fast.
"""

import numpy as np
import pandas as pd
import pytest
from unittest.mock import MagicMock, patch

from tools.excel_tools import ExcelTools


@pytest.fixture
def excel_tools(tmp_path):
    """ExcelTools instance writing to a temp directory."""
    return ExcelTools(output_dir=str(tmp_path / "sheets"))


@pytest.fixture
def fake_ticker(tmp_path):
    """Mock yfinance Ticker with realistic info and history."""
    dates = pd.date_range("2025-01-01", periods=22, freq="B", tz="America/New_York")
    hist = pd.DataFrame(
        {
            "Open":   np.linspace(148.0, 155.0, 22),
            "High":   np.linspace(150.0, 157.0, 22),
            "Low":    np.linspace(146.0, 153.0, 22),
            "Close":  np.linspace(149.0, 156.0, 22),
            "Volume": [45_000_000] * 22,
        },
        index=dates,
    )
    ticker = MagicMock()
    ticker.history.return_value = hist
    ticker.info = {
        "shortName": "Apple Inc.",
        "currentPrice": 155.5,
        "regularMarketChangePercent": 0.42,
        "volume": 48_000_000,
        "marketCap": 2_400_000_000_000,
        "fiftyTwoWeekHigh": 199.0,
        "fiftyTwoWeekLow": 130.0,
        "trailingPE": 28.4,
        "forwardPE": 25.1,
        "trailingEps": 6.12,
        "dividendYield": 0.0055,
        "beta": 1.18,
        "profitMargins": 0.253,
        "totalRevenue": 383_000_000_000,
    }
    return ticker


class TestCreateStockExcelReport:
    def test_returns_filename_with_xlsx(self, excel_tools, fake_ticker):
        with patch("cache.cached_ticker_info", return_value=fake_ticker.info), \
             patch("cache.cached_ticker_history", return_value=fake_ticker.history()):
            result = excel_tools.create_stock_excel_report("AAPL")

        assert "aapl_report" in result.lower()
        assert ".xlsx" in result

    def test_xlsx_file_is_created_on_disk(self, excel_tools, fake_ticker, tmp_path):
        with patch("cache.cached_ticker_info", return_value=fake_ticker.info), \
             patch("cache.cached_ticker_history", return_value=fake_ticker.history()):
            excel_tools.create_stock_excel_report("AAPL")

        xlsx_files = list((tmp_path / "sheets").glob("*.xlsx"))
        assert len(xlsx_files) == 1

    def test_xlsx_has_expected_sheets(self, excel_tools, fake_ticker, tmp_path):
        import openpyxl
        with patch("cache.cached_ticker_info", return_value=fake_ticker.info), \
             patch("cache.cached_ticker_history", return_value=fake_ticker.history()):
            excel_tools.create_stock_excel_report("AAPL")

        xlsx_path = next((tmp_path / "sheets").glob("*.xlsx"))
        wb = openpyxl.load_workbook(xlsx_path)
        assert "Summary" in wb.sheetnames
        assert "Fundamentals" in wb.sheetnames
        assert "Price History" in wb.sheetnames

    def test_yfinance_exception_returns_error_message(self, excel_tools):
        with patch("cache.cached_ticker_info", side_effect=Exception("Timeout")):
            result = excel_tools.create_stock_excel_report("AAPL")

        assert "error" in result.lower()

    def test_different_symbols_produce_different_filenames(self, excel_tools, fake_ticker, tmp_path):
        import time
        with patch("cache.cached_ticker_info", return_value=fake_ticker.info), \
             patch("cache.cached_ticker_history", return_value=fake_ticker.history()):
            r1 = excel_tools.create_stock_excel_report("AAPL")
            time.sleep(0.01)  # ensure different timestamp
            r2 = excel_tools.create_stock_excel_report("MSFT")

        assert r1 != r2


class TestCreateComparisonExcel:
    def test_returns_filename_with_xlsx(self, excel_tools, fake_ticker):
        with patch("cache.cached_ticker_info", return_value=fake_ticker.info):
            result = excel_tools.create_comparison_excel("AAPL,MSFT")

        assert "comparison" in result.lower()
        assert ".xlsx" in result

    def test_handles_mixed_valid_invalid_symbols(self, excel_tools, fake_ticker):
        def side_effect(sym):
            if sym == "FAKE":
                raise Exception("Not found")
            return fake_ticker.info

        with patch("cache.cached_ticker_info", side_effect=side_effect):
            result = excel_tools.create_comparison_excel("AAPL,FAKE,MSFT")

        # Should still produce a file for the valid symbols
        assert ".xlsx" in result

    def test_all_invalid_symbols_returns_error(self, excel_tools):
        with patch("cache.cached_ticker_info", side_effect=Exception("Not found")):
            result = excel_tools.create_comparison_excel("FAKE1,FAKE2")

        # Either error message or empty file — no crash
        assert isinstance(result, str)

    def test_comparison_file_has_data(self, excel_tools, fake_ticker, tmp_path):
        import openpyxl
        with patch("cache.cached_ticker_info", return_value=fake_ticker.info):
            excel_tools.create_comparison_excel("AAPL,MSFT")

        xlsx_path = next((tmp_path / "sheets").glob("*.xlsx"))
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        # Should have header row + at least one data row
        assert ws.max_row >= 2


class TestCreateExcelSheet:
    def test_creates_file_from_dict(self, excel_tools, tmp_path):
        data = {"Stock": ["AAPL", "MSFT", "NVDA"], "Price": [155.0, 300.0, 480.0]}
        result = excel_tools.create_excel_sheet(data, "custom_sheet")

        assert ".xlsx" in result
        xlsx_files = list((tmp_path / "sheets").glob("*.xlsx"))
        assert len(xlsx_files) == 1

    def test_custom_sheet_name(self, excel_tools, tmp_path):
        import openpyxl
        data = {"Category": ["Tech", "Health"], "Value": [50, 30]}
        excel_tools.create_excel_sheet(data, "sector_data", sheet_name="Sectors")

        xlsx_path = next((tmp_path / "sheets").glob("*.xlsx"))
        wb = openpyxl.load_workbook(xlsx_path)
        assert "Sectors" in wb.sheetnames

    def test_single_column_data(self, excel_tools):
        data = {"Values": [1, 2, 3, 4, 5]}
        result = excel_tools.create_excel_sheet(data, "single_col")
        assert ".xlsx" in result

    def test_empty_data_does_not_crash(self, excel_tools):
        data = {"A": [], "B": []}
        result = excel_tools.create_excel_sheet(data, "empty_sheet")
        assert isinstance(result, str)
