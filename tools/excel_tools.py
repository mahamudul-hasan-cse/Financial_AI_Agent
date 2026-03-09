"""Excel/spreadsheet tools for the Financial AI Agent.

Provides tools to create and manipulate Excel spreadsheets
from financial data. Includes self-contained tools that fetch
stock data automatically via yfinance.
"""

from datetime import datetime
from pathlib import Path
from typing import Any

from agno.tools import Toolkit

try:
    import pandas as pd
except ImportError:
    raise ImportError("pandas is required. Install with: pip install pandas")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    raise ImportError("openpyxl is required. Install with: pip install openpyxl")

try:
    import yfinance as yf
except ImportError:
    raise ImportError("yfinance is required. Install with: pip install yfinance")

import cache


class ExcelTools(Toolkit):
    """Toolkit for creating and manipulating Excel spreadsheets."""

    def __init__(self, output_dir: str = "outputs/sheets"):
        """
        Initialize ExcelTools.

        Args:
            output_dir: Directory to save Excel files.
        """
        super().__init__(name="excel_tools")
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.register(self.create_stock_excel_report)
        self.register(self.create_comparison_excel)
        self.register(self.create_excel_sheet)

    def create_stock_excel_report(self, symbol: str) -> str:
        """Create a comprehensive Excel report for a stock by fetching data from Yahoo Finance. Use this when the user asks for a spreadsheet, Excel file, or report for a single stock.

        Args:
            symbol: Stock ticker symbol, e.g. AAPL, GOOGL, MSFT, TSLA.

        Returns:
            A markdown download link to the created Excel file.
        """
        try:
            info = cache.cached_ticker_info(symbol)
            hist = cache.cached_ticker_history(symbol, "1mo")

            filename = f"{symbol.lower()}_report_{self._timestamp()}"
            filepath = self.output_dir / f"{filename}.xlsx"

            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                # Summary sheet
                summary = {
                    "Symbol": [symbol.upper()],
                    "Name": [info.get("shortName", "N/A")],
                    "Price": [info.get("currentPrice") or info.get("regularMarketPrice", "N/A")],
                    "Day Change (%)": [info.get("regularMarketChangePercent", "N/A")],
                    "Volume": [info.get("volume", "N/A")],
                    "Market Cap": [info.get("marketCap", "N/A")],
                    "52W High": [info.get("fiftyTwoWeekHigh", "N/A")],
                    "52W Low": [info.get("fiftyTwoWeekLow", "N/A")],
                }
                pd.DataFrame(summary).to_excel(writer, sheet_name="Summary", index=False)

                # Fundamentals sheet
                fundamentals = {
                    "P/E Ratio": [info.get("trailingPE", "N/A")],
                    "Forward P/E": [info.get("forwardPE", "N/A")],
                    "EPS": [info.get("trailingEps", "N/A")],
                    "Dividend Yield": [info.get("dividendYield", "N/A")],
                    "Beta": [info.get("beta", "N/A")],
                    "Profit Margin": [info.get("profitMargins", "N/A")],
                    "Revenue": [info.get("totalRevenue", "N/A")],
                }
                pd.DataFrame(fundamentals).to_excel(writer, sheet_name="Fundamentals", index=False)

                # Price history sheet
                if not hist.empty:
                    price_hist = hist[["Open", "High", "Low", "Close", "Volume"]].copy()
                    price_hist.index = price_hist.index.strftime("%Y-%m-%d")
                    price_hist.index.name = "Date"
                    price_hist.to_excel(writer, sheet_name="Price History")

            self._apply_workbook_styling(filepath)
            return f"Excel report for {symbol.upper()} saved as {filename}.xlsx"
        except Exception as e:
            return f"Error creating Excel report for {symbol}: {e}"

    def create_comparison_excel(self, symbols: str) -> str:
        """Create an Excel spreadsheet comparing multiple stocks by fetching data from Yahoo Finance. Use this when the user asks to compare stocks in a spreadsheet or Excel file.

        Args:
            symbols: Comma-separated stock ticker symbols, e.g. 'AAPL,GOOGL,MSFT' or 'AAPL, GOOGL, MSFT'.

        Returns:
            A markdown download link to the created Excel file.
        """
        try:
            symbol_list = [s.strip().upper() for s in symbols.split(",") if s.strip()]
            rows = []

            for sym in symbol_list:
                try:
                    info = cache.cached_ticker_info(sym)
                    rows.append({
                        "Symbol": sym,
                        "Name": info.get("shortName", "N/A"),
                        "Price ($)": info.get("currentPrice") or info.get("regularMarketPrice", "N/A"),
                        "Change (%)": info.get("regularMarketChangePercent", "N/A"),
                        "Market Cap": info.get("marketCap", "N/A"),
                        "P/E Ratio": info.get("trailingPE", "N/A"),
                        "EPS": info.get("trailingEps", "N/A"),
                        "52W High": info.get("fiftyTwoWeekHigh", "N/A"),
                        "52W Low": info.get("fiftyTwoWeekLow", "N/A"),
                    })
                except Exception:
                    rows.append({"Symbol": sym, "Name": "Error fetching data"})

            if not rows:
                return "Could not fetch data for any of the provided symbols."

            df = pd.DataFrame(rows)
            filename = f"stock_comparison_{self._timestamp()}"
            filepath = self.output_dir / f"{filename}.xlsx"

            return self._create_styled_excel(df, filepath, "Stock Comparison")
        except Exception as e:
            return f"Error creating comparison sheet: {e}"

    def create_excel_sheet(
        self,
        data: dict,
        filename: str,
        sheet_name: str = "Sheet1",
    ) -> str:
        """Create an Excel file from dictionary data. Use this for creating custom spreadsheets with arbitrary data.

        Args:
            data: Dictionary with column names as keys and lists as values. Example: {"Stock": ["AAPL", "GOOGL"], "Price": [150.0, 140.0]}
            filename: Name for the output file without extension.
            sheet_name: Name of the sheet in the Excel file.

        Returns:
            A markdown download link to the created Excel file.
        """
        try:
            df = pd.DataFrame(data)
            filepath = self.output_dir / f"{filename}.xlsx"
            return self._create_styled_excel(df, filepath, sheet_name)
        except Exception as e:
            return f"Error creating Excel file: {e}"

    def _create_styled_excel(
        self,
        df: pd.DataFrame,
        filepath: Path,
        sheet_name: str,
    ) -> str:
        """Create a styled Excel file from a DataFrame."""
        df.to_excel(filepath, sheet_name=sheet_name, index=False)

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)
        wb.close()

        return f"Excel file saved as {filepath.name}"

    def _apply_workbook_styling(self, filepath: Path) -> None:
        """Apply professional styling to all sheets in a workbook."""
        wb = openpyxl.load_workbook(filepath)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except TypeError:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)
        wb.close()

    def _timestamp(self) -> str:
        """Generate a timestamp string for unique filenames."""
        return datetime.now().strftime("%Y%m%d_%H%M%S")
